# -*- coding: utf-8 -*-

#Created on Sun Jul 22 09:18:30 2018

#@author: akaliontzakis

from operator import itemgetter
import openpyxl, smtplib, sys
from datetime import datetime
from datetime import date
from email.mime.text import MIMEText
import pprint
from random import seed
from random import randint


def maketitles():

    #first roll to determine overall rarity. second roll to determine rarity of the beginning of title, if applicable. Third roll to determine rarity of the end of title, if applicable
    roll_1 = randint(0, 100)
    roll_2 = randint(0, 100)
    roll_3 = randint(0, 100)
    titlelist = []
    finaltitles = []
    r1rarity = ''
    r2rarity = ''
    r3rarity = ''

    #list of excel workbooks to iterate through
    workbook = './titles.xlsx'
    #create list of tuples


    #iterate through workbooks

    wb = openpyxl.load_workbook(workbook)
        #select sheet
    sheet = wb['Sheet1']

    #iterate thru sheet
    for r in range(2, sheet.max_row + 1):
        title = sheet.cell(row=r, column=1).value
        location = sheet.cell(row=r, column=2).value
        rarity = sheet.cell(row=r, column=3).value
        titlelist.append((title,location,rarity))
        

    length = len(titlelist) - 1

    #low rarity title
    if roll_1 < 61:
        r1rarity = 'low'
        lowraritytitle = titlelist[randint(1,length)]
        print (lowraritytitle)
        while True:
            if (lowraritytitle[1] == "Middle") and (lowraritytitle[2] == "Low"):
                finaltitles.append((r1rarity, lowraritytitle[0]))
                break
            lowraritytitle = titlelist[randint(1,length)]

        
    #medium rarity title   
    if roll_1 in range(61,95):

        r1rarity = 'medium'
        medraritytitle = ""
        medraritytitlebeg = titlelist[randint(1,length)]
        medraritytitlemid = titlelist[randint(1,length)]
        #determine the beginning of the title
        if roll_2 < 61:
            r2rarity = 'low'
            while True:
                if (medraritytitlebeg[2] == "Low") and (medraritytitlebeg[1] == "Beginning"):
                    medraritytitle = medraritytitlebeg[0]
                    break
                medraritytitlebeg = titlelist[randint(1,length)]
                
        if roll_2 in range (61,95):
            r2rarity = 'medium'
            while True:
                if (medraritytitlebeg[2] == "Medium") and (medraritytitlebeg[1] == "Beginning"):
                    medraritytitle = medraritytitlebeg[0]
                    break
                medraritytitlebeg = titlelist[randint(1,length)]
            
        if roll_2 in range (96,100):
            r2rarity = 'high'
            while True:
                if (medraritytitlebeg[2] == "Medium") and (medraritytitlebeg[1] == "Beginning"):
                    medraritytitle = medraritytitlebeg[0]  
                    break
                medraritytitlebeg = titlelist[randint(1,length)]
               
            
            
        #determine the end of the title        
        while True:
            if (medraritytitlemid[2] == "Medium")  and (medraritytitlemid[1] == "Middle"):
                medraritytitle = medraritytitle + " " + medraritytitlemid[0]
                break
            medraritytitlemid = titlelist[randint(1,length)]
        

        finaltitles.append((r1rarity, r2rarity, medraritytitle))    

    #high rarity title
    if roll_1 in range (96,100):

        r1rarity = 'high'    
        highraritytitle = ""
        highraritytitlebeg = titlelist[randint(1,length)]
        highraritytitlemid = titlelist[randint(1,length)]
        highraritytitleend = titlelist[randint(1,length)]
            #determine the beginning of the title
        if roll_2 < 61:
            r2rarity = 'low'
            while True:
                if (highraritytitlebeg[2] == "Low") and (highraritytitlebeg[1] == "Beginning"):
                    highraritytitle = highraritytitlebeg[0]
                    break  
                highraritytitlebeg = titlelist[randint(1,length)]
            
        if roll_2 in range (61,95):
            r2rarity = 'medium'
            while True:
                if (highraritytitlebeg[2] == "Medium") and (highraritytitlebeg[1] == "Beginning"):
                    highraritytitle = highraritytitlebeg[0]
                    break
                highraritytitlebeg = titlelist[randint(1,length)]
            
        if roll_2 in range (96,100):
            r2rarity = 'high'
            while True:
                if (highraritytitlebeg[2] == "High") and (highraritytitlebeg[1] == "Beginning"):
                    highraritytitle = highraritytitlebeg[0]   
                    break
                highraritytitlebeg = titlelist[randint(1,length)]
              
      
        #determine middle of title
        while True:
            if (highraritytitlemid[2] == "High") and (highraritytitlemid[1] == "Middle"):
                highraritytitle = highraritytitle + " " + highraritytitlemid[0] 
                break
            highraritytitlemid = titlelist[randint(1,length)]
             
          
        #determine the end of the title      
        if roll_3 < 61:
            r3rarity = 'low'  
            while True:
                if (highraritytitleend[2] == "Low") and (highraritytitleend[1] == "End"):
                    highraritytitle = highraritytitle + " " + highraritytitleend[0]
                    break
                highraritytitleend = titlelist[randint(1,length)]
             
        if roll_3 in range (61,95):
            r3rarity = 'medium'  
            while True:
                if (highraritytitleend[2] == "Medium")  and (highraritytitleend[1] == "End"):
                    highraritytitle = highraritytitle + " " + highraritytitleend[0]
                    break
                highraritytitleend = titlelist[randint(1,length)]
            
        if roll_3 in range (96,100):
            r3rarity = 'high'  
            while True:
                if (highraritytitleend[2] == "High")  and (highraritytitleend[1] == "End"):
                    highraritytitle = highraritytitle + " " + highraritytitleend[0] 
                    break
                highraritytitleend = titlelist[randint(1,length)]
               

        finaltitles.append((r1rarity, r2rarity, r3rarity, highraritytitle))  


    print(finaltitles)
#    for row in finaltitles:
 #       print ('%s | %s | %s | %s \n' % row)

    
for i in range(0,10):
    maketitles()